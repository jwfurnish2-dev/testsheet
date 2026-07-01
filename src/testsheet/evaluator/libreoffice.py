"""
M1 Spike — Evaluator B: LibreOffice headless recalculator.

Pros:  near-100% Excel formula compatibility; handles volatile functions.
Cons:  requires LibreOffice installed as a system dependency.

Install LibreOffice:
  Ubuntu/Debian: apt-get install libreoffice
  macOS:         brew install --cask libreoffice
  Windows:       https://www.libreoffice.org/download/
  GitHub Actions: see action.yml

Usage:
  testsheet baseline model.xlsx --evaluator libreoffice
"""

from __future__ import annotations

import logging
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from testsheet.evaluator.base import BaseEvaluator

logger = logging.getLogger(__name__)

# Default paths to search for the soffice binary
_SOFFICE_CANDIDATES = [
    "soffice",
    "/usr/bin/soffice",
    "/usr/lib/libreoffice/program/soffice",
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
]


def _find_soffice() -> str | None:
    for candidate in _SOFFICE_CANDIDATES:
        if shutil.which(candidate):
            return candidate
        p = Path(candidate)
        if p.exists():
            return str(p)
    return None


class LibreOfficeEvaluator(BaseEvaluator):
    """
    Evaluate a workbook by running LibreOffice headless to force recalculation,
    then reading the recalculated values with openpyxl.
    """

    def __init__(self, soffice_path: str | None = None) -> None:
        self._soffice = soffice_path or _find_soffice()

    @property
    def name(self) -> str:
        return "libreoffice"

    def is_available(self) -> bool:
        return self._soffice is not None

    def compute(
        self,
        path: Path,
        sheets: list[str] | None = None,
    ) -> dict[str, dict[str, Any]]:
        """
        Recalculate *path* via LibreOffice headless and return computed values.

        Strategy
        --------
        1. Copy the workbook to a temp dir.
        2. Call ``soffice --headless --convert-to xlsx`` — LibreOffice opens
           the file, recalculates all cells, and writes a new xlsx.
        3. Read the recalculated file with openpyxl ``data_only=True``.
        """
        if not self.is_available():
            logger.warning(
                "LibreOffice (soffice) not found on PATH — returning empty value map. "
                "Install LibreOffice or switch to --evaluator pycel."
            )
            return {}

        path = Path(path).resolve()

        with tempfile.TemporaryDirectory(prefix="testsheet_lo_") as tmp:
            tmp_path = Path(tmp)
            # Copy workbook into the temp dir so LO output lands there too
            import shutil as _shutil
            tmp_wb = tmp_path / path.name
            _shutil.copy2(path, tmp_wb)

            cmd = [
                self._soffice,
                "--headless",
                "--norestore",
                "--convert-to", "xlsx",
                "--outdir", str(tmp_path),
                str(tmp_wb),
            ]

            logger.debug("LibreOfficeEvaluator: %s", " ".join(cmd))

            try:
                proc = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
            except FileNotFoundError as exc:
                logger.error("soffice not found: %s", exc)
                return {}
            except subprocess.TimeoutExpired:
                logger.error("LibreOffice timed out after 120 s")
                return {}

            if proc.returncode != 0:
                logger.error("LibreOffice exited %d: %s", proc.returncode, proc.stderr)
                return {}

            # LO produces a file with the same stem; find it
            recalced = tmp_path / path.name
            if not recalced.exists():
                # Sometimes LO appends nothing, sometimes something; glob
                candidates = list(tmp_path.glob("*.xlsx"))
                candidates = [c for c in candidates if c != tmp_wb]
                if not candidates:
                    logger.error("LibreOffice produced no xlsx output in %s", tmp_path)
                    return {}
                recalced = candidates[0]

            return _extract_values(recalced, sheets=sheets)


def _extract_values(
    path: Path,
    sheets: list[str] | None = None,
) -> dict[str, dict[str, Any]]:
    """Read openpyxl data_only values from a recalculated workbook."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result: dict[str, dict[str, Any]] = {}

    for sheet_name in (sheets or wb.sheetnames):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_out: dict[str, Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                addr = f"{get_column_letter(cell.column)}{cell.row}"
                sheet_out[addr] = cell.value
        result[sheet_name] = sheet_out

    wb.close()
    return result
