"""
M1 Spike — Evaluator A: pure-Python pycel library.

Pros:  zero system deps, CI-friendly, fast install.
Cons:  ~80% formula coverage; XLOOKUP, dynamic arrays, volatile functions
       (NOW, RAND) not supported. Cross-sheet references supported.

Install:  pip install "testsheet[pycel]"

pycel API notes (v1.0b30+):
  - ExcelCompiler(filename=...) loads and compiles the workbook graph
  - compiler.evaluate("Sheet!A1") evaluates a single cell by reference
  - No evaluate_all() — must iterate cells and call evaluate() per address
  - Errors surface as pycel.excelutil.ERROR types; we convert to Excel strings
"""

from __future__ import annotations

import logging
import math
from pathlib import Path
from typing import Any

from testsheet.evaluator.base import BaseEvaluator

logger = logging.getLogger(__name__)

# Map pycel error tokens → Excel error strings
_PYCEL_ERROR_MAP = {
    "DIV/0": "#DIV/0!",
    "N/A":   "#N/A",
    "NAME":  "#NAME?",
    "NULL":  "#NULL!",
    "NUM":   "#NUM!",
    "REF":   "#REF!",
    "VALUE": "#VALUE!",
}


def _convert_pycel_value(val: Any) -> Any:
    """Convert pycel output to a plain Python value."""
    if val is None:
        return None
    # pycel error objects have a .value or str like "DIV/0"
    val_str = str(val)
    for token, excel_err in _PYCEL_ERROR_MAP.items():
        if token in val_str:
            return excel_err
    # numpy scalars → native Python
    try:
        import numpy as np
        if isinstance(val, np.generic):
            return val.item()
    except ImportError:
        pass
    # Float NaN/Inf → None (treat as empty)
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    return val


class PycelEvaluator(BaseEvaluator):
    """Evaluate workbook using the pycel library (pure Python)."""

    @property
    def name(self) -> str:
        return "pycel"

    def is_available(self) -> bool:
        try:
            import pycel  # noqa: F401
            return True
        except ImportError:
            return False

    def compute(
        self,
        path: Path,
        sheets: list[str] | None = None,
    ) -> dict[str, dict[str, Any]]:
        """
        Compute cell values via pycel's ExcelCompiler.

        Returns ``{sheet_name: {address: value}}``.
        Cells that pycel cannot evaluate are omitted — the parser falls back
        to openpyxl cached values for those.
        """
        if not self.is_available():
            logger.warning(
                "pycel not installed — returning empty value map (openpyxl cached values used). "
                "Run: pip install 'testsheet[pycel]'"
            )
            return {}

        path = Path(path)
        result: dict[str, dict[str, Any]] = {}

        try:
            from pycel import ExcelCompiler  # type: ignore[import]
        except ImportError:
            return {}

        try:
            compiler = ExcelCompiler(filename=str(path))
        except Exception as exc:
            logger.error("PycelEvaluator: failed to load %s: %s", path.name, exc)
            return {}

        # Determine which sheets + addresses to evaluate
        import openpyxl
        from openpyxl.utils import get_column_letter

        try:
            wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        except Exception as exc:
            logger.error("PycelEvaluator: openpyxl failed to open %s: %s", path.name, exc)
            return {}

        target_sheets = sheets or wb.sheetnames

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_out: dict[str, Any] = {}

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    addr = f"{get_column_letter(cell.column)}{cell.row}"
                    ref = f"{sheet_name}!{addr}"
                    try:
                        raw = compiler.evaluate(ref)
                        sheet_out[addr] = _convert_pycel_value(raw)
                    except Exception as exc:
                        # pycel can't evaluate this cell — leave it out,
                        # parser will use openpyxl cached value instead
                        logger.debug("pycel could not evaluate %s: %s", ref, exc)

            result[sheet_name] = sheet_out

        wb.close()
        return result
