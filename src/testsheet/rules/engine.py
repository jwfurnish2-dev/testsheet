"""
Invariant rule engine — load rules.yaml and evaluate each rule against
the current workbook.

Rule types (MVP):
  range_bound          — cell/range value within [min, max]
  no_error             — no Excel error string in range
  no_hardcode_in_range — every cell in range must be a formula (not a literal)
  totals_tie           — total_cell == SUM(sum_range) within tolerance
  monotonic            — values in range are monotonically increasing/decreasing
  relationship         — arbitrary cross-cell expression evaluates to True
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from testsheet.evaluator.base import BaseEvaluator


# ── Data classes ────────────────────────────────────────────────────────────

@dataclass
class RuleResult:
    rule_id: str
    rule_type: str
    passed: bool
    message: str
    detail: dict = field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            "rule_id": self.rule_id,
            "rule_type": self.rule_type,
            "passed": self.passed,
            "message": self.message,
            "detail": self.detail,
        }


# ── YAML loader ─────────────────────────────────────────────────────────────

def load_rules(rules_path: Path) -> list[dict]:
    """Parse rules.yaml and return list of raw rule dicts."""
    data = yaml.safe_load(Path(rules_path).read_text(encoding="utf-8"))
    return data.get("rules", []) if data else []


# ── Rule evaluation entry point ─────────────────────────────────────────────

def evaluate_rules(
    workbook_path: Path,
    rules: list[dict],
    evaluator: BaseEvaluator,
) -> list[RuleResult]:
    """Evaluate all *rules* against *workbook_path* and return results."""
    if not rules:
        return []

    workbook_path = Path(workbook_path).resolve()
    computed = evaluator.compute(workbook_path)

    # Load formulas separately for no_hardcode check
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)

    results: list[RuleResult] = []
    for rule in rules:
        result = _evaluate_one(rule, workbook_path, computed, wb_formulas)
        results.append(result)

    wb_formulas.close()
    return results


# ── Individual rule handlers ─────────────────────────────────────────────────

def _evaluate_one(
    rule: dict,
    workbook_path: Path,
    computed: dict[str, dict[str, Any]],
    wb_formulas,
) -> RuleResult:
    rule_id = rule.get("id", "unnamed")
    rule_type = rule.get("type", "")

    handlers = {
        "range_bound": _rule_range_bound,
        "no_error": _rule_no_error,
        "no_hardcode_in_range": _rule_no_hardcode,
        "totals_tie": _rule_totals_tie,
        "monotonic": _rule_monotonic,
        "relationship": _rule_relationship,
    }

    handler = handlers.get(rule_type)
    if handler is None:
        return RuleResult(
            rule_id=rule_id,
            rule_type=rule_type,
            passed=False,
            message=f"Unknown rule type '{rule_type}'",
        )

    try:
        return handler(rule, computed, wb_formulas)
    except Exception as exc:
        return RuleResult(
            rule_id=rule_id,
            rule_type=rule_type,
            passed=False,
            message=f"Rule evaluation error: {exc}",
        )


def _rule_range_bound(rule: dict, computed: dict, _wb) -> RuleResult:
    rule_id = rule["id"]
    ref = rule.get("cell") or rule.get("range")
    min_val = rule.get("min")
    max_val = rule.get("max")
    tol = rule.get("tolerance", 1e-9)

    cells = _resolve_ref(ref, computed)
    failures = []
    for (sheet, addr, value) in cells:
        if value is None:
            continue
        try:
            v = float(value)
        except (TypeError, ValueError):
            failures.append(f"{sheet}!{addr}={value!r} (not numeric)")
            continue
        if min_val is not None and v < min_val - tol:
            failures.append(f"{sheet}!{addr}={v} < min={min_val}")
        if max_val is not None and v > max_val + tol:
            failures.append(f"{sheet}!{addr}={v} > max={max_val}")

    passed = not failures
    return RuleResult(
        rule_id=rule_id,
        rule_type="range_bound",
        passed=passed,
        message="OK" if passed else "; ".join(failures),
        detail={"failures": failures},
    )


_EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}


def _rule_no_error(rule: dict, computed: dict, _wb) -> RuleResult:
    rule_id = rule["id"]
    ref = rule.get("range") or rule.get("cell")
    cells = _resolve_ref(ref, computed)

    errors = [
        f"{sheet}!{addr}={value!r}"
        for (sheet, addr, value) in cells
        if isinstance(value, str) and value.strip().upper() in _EXCEL_ERRORS
    ]
    passed = not errors
    return RuleResult(
        rule_id=rule_id,
        rule_type="no_error",
        passed=passed,
        message="OK" if passed else f"Error cells: {errors}",
        detail={"error_cells": errors},
    )


def _rule_no_hardcode(rule: dict, computed: dict, wb_formulas) -> RuleResult:
    rule_id = rule["id"]
    ref = rule.get("range") or rule.get("cell")
    sheet_name, range_str = _split_ref(ref)

    ws = wb_formulas[sheet_name] if sheet_name in wb_formulas.sheetnames else None
    if ws is None:
        return RuleResult(rule_id=rule_id, rule_type="no_hardcode_in_range",
                          passed=False, message=f"Sheet '{sheet_name}' not found")

    hardcoded = []
    for addr in _range_addresses(range_str):
        cell = ws[addr]
        raw = cell.value
        if raw is not None and not (isinstance(raw, str) and raw.startswith("=")):
            hardcoded.append(f"{sheet_name}!{addr}={raw!r}")

    passed = not hardcoded
    return RuleResult(
        rule_id=rule_id,
        rule_type="no_hardcode_in_range",
        passed=passed,
        message="OK" if passed else f"Hardcoded cells: {hardcoded}",
        detail={"hardcoded_cells": hardcoded},
    )


def _rule_totals_tie(rule: dict, computed: dict, _wb) -> RuleResult:
    rule_id = rule["id"]
    total_ref = rule["total_cell"]
    sum_ref = rule["sum_range"]
    tol = rule.get("tolerance", 1e-9)

    total_cells = _resolve_ref(total_ref, computed)
    if not total_cells:
        return RuleResult(rule_id=rule_id, rule_type="totals_tie",
                          passed=False, message=f"Total cell not found: {total_ref}")
    _, _, total_val = total_cells[0]

    sum_cells = _resolve_ref(sum_ref, computed)
    computed_sum = sum(
        float(v) for _, _, v in sum_cells
        if v is not None and not isinstance(v, str)
    )

    try:
        total_float = float(total_val)
    except (TypeError, ValueError):
        return RuleResult(rule_id=rule_id, rule_type="totals_tie",
                          passed=False, message=f"Total cell value is not numeric: {total_val!r}")

    passed = math.isclose(total_float, computed_sum, rel_tol=tol, abs_tol=tol)
    msg = "OK" if passed else f"total={total_float} != sum={computed_sum} (diff={abs(total_float-computed_sum):.6g})"
    return RuleResult(rule_id=rule_id, rule_type="totals_tie", passed=passed, message=msg,
                      detail={"total": total_float, "computed_sum": computed_sum})


def _rule_monotonic(rule: dict, computed: dict, _wb) -> RuleResult:
    rule_id = rule["id"]
    ref = rule.get("range")
    direction = rule.get("direction", "increasing")  # "increasing" | "decreasing"
    strict = rule.get("strict", False)

    cells = _resolve_ref(ref, computed)
    values = [float(v) for _, _, v in cells if v is not None and not isinstance(v, str)]

    violations = []
    for i in range(len(values) - 1):
        a, b = values[i], values[i + 1]
        if direction == "increasing":
            ok = (b > a) if strict else (b >= a)
        else:
            ok = (b < a) if strict else (b <= a)
        if not ok:
            violations.append(f"index {i}: {a} -> {b}")

    passed = not violations
    return RuleResult(rule_id=rule_id, rule_type="monotonic",
                      passed=passed,
                      message="OK" if passed else f"Monotonic violations: {violations[:5]}",
                      detail={"violations": violations})


def _rule_relationship(rule: dict, computed: dict, _wb) -> RuleResult:
    """
    Evaluate a simple cross-cell expression.

    Example rule:
      type: relationship
      expression: "Summary!B10 >= Summary!B9 * 0.9"
    """
    rule_id = rule["id"]
    expr = rule.get("expression", "")

    # Build a namespace of {sheet_addr: value} for safe eval
    namespace: dict[str, Any] = {}
    # Replace "Sheet!A1" references with Python variable names
    pattern = re.compile(r"([A-Za-z0-9_ ]+)!([A-Z]+[0-9]+)")

    def replace_ref(m: re.Match) -> str:
        sheet, addr = m.group(1).strip(), m.group(2)
        var = f"_ref_{re.sub(r'[^A-Za-z0-9]', '_', sheet)}_{addr}"
        val = (computed.get(sheet) or {}).get(addr)
        namespace[var] = float(val) if val is not None else None
        return var

    safe_expr = pattern.sub(replace_ref, expr)

    try:
        result = eval(safe_expr, {"__builtins__": {}}, namespace)  # noqa: S307
        passed = bool(result)
    except Exception as exc:
        return RuleResult(rule_id=rule_id, rule_type="relationship",
                          passed=False, message=f"Expression error: {exc}",
                          detail={"expression": expr})

    return RuleResult(rule_id=rule_id, rule_type="relationship",
                      passed=passed,
                      message="OK" if passed else f"Expression false: {expr}",
                      detail={"expression": expr, "namespace": {k: v for k, v in namespace.items()}})


# ── Helpers ──────────────────────────────────────────────────────────────────

def _split_ref(ref: str) -> tuple[str, str]:
    """Split 'Sheet!A1:B5' into ('Sheet', 'A1:B5') or ('Sheet', 'A1')."""
    if "!" in ref:
        sheet, addr = ref.split("!", 1)
        return sheet, addr
    # No sheet prefix — ambiguous; return as-is with empty sheet
    return "", ref


def _range_addresses(range_str: str) -> list[str]:
    """Expand 'A1:C3' into ['A1','A2','A3','B1',...,'C3']."""
    if ":" not in range_str:
        return [range_str]
    start, end = range_str.split(":")
    col_s = re.match(r"([A-Z]+)", start).group(1)
    row_s = int(re.search(r"(\d+)", start).group(1))
    col_e = re.match(r"([A-Z]+)", end).group(1)
    row_e = int(re.search(r"(\d+)", end).group(1))

    col_start = column_index_from_string(col_s)
    col_end = column_index_from_string(col_e)

    addrs = []
    for c in range(col_start, col_end + 1):
        for r in range(row_s, row_e + 1):
            addrs.append(f"{get_column_letter(c)}{r}")
    return addrs


def _resolve_ref(
    ref: str,
    computed: dict[str, dict[str, Any]],
) -> list[tuple[str, str, Any]]:
    """
    Resolve a ref like 'Sheet!A1' or 'Sheet!A1:B5' to a list of
    (sheet_name, address, value) triples.
    """
    sheet, range_str = _split_ref(ref)
    addrs = _range_addresses(range_str)
    sheet_vals = computed.get(sheet, {})
    return [(sheet, a, sheet_vals.get(a)) for a in addrs]
