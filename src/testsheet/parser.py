"""Workbook parser — reads cell values and formulas via openpyxl."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class CellSnapshot:
    """Single-cell snapshot: address, raw formula string, cached/computed value."""
    sheet: str
    address: str          # e.g. "B5"
    formula: str | None   # "=SUM(A1:A4)" or None if literal
    value: Any            # computed / cached value
    data_type: str = ""   # openpyxl data_type character


@dataclass
class WorkbookSnapshot:
    """Full snapshot of a workbook's cell map + named ranges."""
    path: Path
    cells: dict[str, dict[str, CellSnapshot]] = field(default_factory=dict)
    # {sheet_name: {address: CellSnapshot}}
    named_ranges: dict[str, str] = field(default_factory=dict)
    # {name: "Sheet!A1:B5"}


def parse_workbook(
    path: Path,
    sheets: list[str] | None = None,
    *,
    computed_values: dict[str, dict[str, Any]] | None = None,
) -> WorkbookSnapshot:
    """
    Parse *path* and return a WorkbookSnapshot.

    Parameters
    ----------
    path:
        Workbook to parse.
    sheets:
        If given, restrict to these sheet names.
    computed_values:
        Optional pre-computed value map ``{sheet: {address: value}}`` provided
        by an Evaluator.  When supplied these override openpyxl's cached values
        (which may be stale).  Structure must match the cell map produced by
        this function.
    """
    path = Path(path)

    # Read formulas (data_only=False)
    wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    # Read cached values (data_only=True) as fallback when no evaluator
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)

    snapshot = WorkbookSnapshot(path=path)

    target_sheets = sheets or wb_formulas.sheetnames

    for sheet_name in target_sheets:
        if sheet_name not in wb_formulas.sheetnames:
            continue

        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        sheet_cells: dict[str, CellSnapshot] = {}

        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                if cell_f.value is None and cell_v.value is None:
                    continue  # skip truly empty cells

                address = f"{get_column_letter(cell_f.column)}{cell_f.row}"
                raw = cell_f.value

                if isinstance(raw, str) and raw.startswith("="):
                    formula: str | None = raw
                    cached_value: Any = cell_v.value
                else:
                    formula = None
                    cached_value = raw  # literal

                # Override with evaluator-computed value if available
                ev_value = (
                    (computed_values or {}).get(sheet_name, {}).get(address)
                )
                final_value = ev_value if ev_value is not None else cached_value

                sheet_cells[address] = CellSnapshot(
                    sheet=sheet_name,
                    address=address,
                    formula=formula,
                    value=final_value,
                    data_type=cell_f.data_type or "",
                )

        snapshot.cells[sheet_name] = sheet_cells

    # Named ranges
    for name, named_range in wb_formulas.defined_names.items():
        try:
            snapshot.named_ranges[name] = str(named_range.attr_text)
        except Exception:
            pass

    wb_formulas.close()
    wb_values.close()

    return snapshot
