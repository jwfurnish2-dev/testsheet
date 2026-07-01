"""
M2 diff engine tests — acceptance criteria for F1 and F2.

Uses FormulasEvaluator (the M1-selected evaluator) so computed values
are real — not openpyxl cached stubs.

F1 AC: Re-running against an unchanged file → PASS, zero drift.
F2 AC: Changing one formula result → exactly one drift, correctly classified,
        with sheet + address + before/after values.
F2 AC: Introducing a #DIV/0! → flagged as error_introduced.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures"
sys.path.insert(0, str(FIXTURES_DIR))


# ── Skip entire module if formulas not installed ──────────────────────────────

formulas = pytest.importorskip(
    "formulas",
    reason="formulas not installed — run: pip install formulas",
)


# ── Module-level fixtures ─────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def ev():
    from testsheet.evaluator.formulas_eval import FormulasEvaluator
    return FormulasEvaluator()


@pytest.fixture(scope="module")
def drift_fixtures():
    from make_drift_fixtures import build_all_drift
    return build_all_drift(FIXTURES_DIR)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _run_diff(before_path: Path, after_path: Path, ev) -> list:
    from testsheet.baseline import capture_baseline, load_baseline
    from testsheet.diff import diff_workbook

    baseline_path = capture_baseline(before_path, evaluator=ev)
    baseline = load_baseline(baseline_path)
    return diff_workbook(after_path, baseline, evaluator=ev)


def _drift_for(drifts: list, sheet: str, address: str):
    """Return the single drift matching sheet+address, or None."""
    matches = [d for d in drifts if d.sheet == sheet and d.address == address]
    return matches[0] if matches else None


# ── F1: Zero drift on same file ───────────────────────────────────────────────

class TestF1ZeroDrift:
    def test_no_drift_literal_workbook(self, tmp_path, ev):
        """Baseline a simple literal workbook, run against itself → zero drift."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws["A1"] = 42
        ws["A2"] = 100
        path = tmp_path / "same.xlsx"
        wb.save(path)

        drifts = _run_diff(path, path, ev)
        assert drifts == [], f"Expected zero drift on same file, got: {drifts}"

    def test_no_drift_formula_workbook(self, tmp_path, ev):
        """Baseline a formula workbook, run against itself → zero drift."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["B1"] = "=A1+A2"
        ws["B2"] = "=SUM(A1:A2)*2"
        path = tmp_path / "formula_same.xlsx"
        wb.save(path)

        drifts = _run_diff(path, path, ev)
        assert drifts == [], f"Expected zero drift on same formula file, got: {drifts}"

    def test_no_drift_m1_simple_model(self, ev):
        """The M1 simple model baseline vs itself → zero drift."""
        path = FIXTURES_DIR / "model_simple.xlsx"
        if not path.exists():
            pytest.skip("M1 fixtures not built yet")
        drifts = _run_diff(path, path, ev)
        assert drifts == []


# ── F2: Drift classification ──────────────────────────────────────────────────

class TestF2DriftClassification:

    def test_value_only_drift(self, drift_fixtures, ev):
        """Literal input change → value_only drift on input cell."""
        before = FIXTURES_DIR / "value_only_before.xlsx"
        after  = FIXTURES_DIR / "value_only_after.xlsx"
        drifts = _run_diff(before, after, ev)

        kinds = {d.address: d.kind for d in drifts if d.sheet == "Sheet"}
        # A1 changed literal: 10 → 99
        assert kinds.get("A1") == "value_only", f"A1 kind={kinds.get('A1')!r}"
        # B1 = A1*10 → value changed, formula unchanged
        assert kinds.get("B1") == "value_only", f"B1 kind={kinds.get('B1')!r}"

    def test_value_only_before_after(self, drift_fixtures, ev):
        """value_only drift captures correct before/after values."""
        before = FIXTURES_DIR / "value_only_before.xlsx"
        after  = FIXTURES_DIR / "value_only_after.xlsx"
        drifts = _run_diff(before, after, ev)

        d = _drift_for(drifts, "Sheet", "A1")
        assert d is not None, "Expected drift on Sheet!A1"
        assert d.baseline_value == 10
        assert d.current_value  == 99

    def test_formula_only_drift(self, drift_fixtures, ev):
        """Formula string change with same result → formula_only."""
        before = FIXTURES_DIR / "formula_only_before.xlsx"
        after  = FIXTURES_DIR / "formula_only_after.xlsx"
        drifts = _run_diff(before, after, ev)

        b1 = _drift_for(drifts, "Sheet", "B1")
        assert b1 is not None, "Expected drift on Sheet!B1"
        assert b1.kind == "formula_only", f"Expected formula_only, got {b1.kind!r}"
        assert b1.baseline_formula == "=A1+A2+A3"
        assert b1.current_formula  == "=SUM(A1:A3)"

    def test_both_drift(self, drift_fixtures, ev):
        """Formula change with different result → both."""
        before = FIXTURES_DIR / "both_before.xlsx"
        after  = FIXTURES_DIR / "both_after.xlsx"
        drifts = _run_diff(before, after, ev)

        b1 = _drift_for(drifts, "Sheet", "B1")
        assert b1 is not None, "Expected drift on Sheet!B1"
        assert b1.kind == "both", f"Expected both, got {b1.kind!r}"
        assert b1.baseline_value == 20
        assert b1.current_value  == 110

    def test_error_introduced(self, drift_fixtures, ev):
        """Division-by-zero introduced → error_introduced classification."""
        before = FIXTURES_DIR / "error_before.xlsx"
        after  = FIXTURES_DIR / "error_after.xlsx"
        drifts = _run_diff(before, after, ev)

        b1 = _drift_for(drifts, "Sheet", "B1")
        assert b1 is not None, "Expected drift on Sheet!B1"
        assert b1.kind == "error_introduced", f"Expected error_introduced, got {b1.kind!r}"

    def test_new_cell(self, drift_fixtures, ev):
        """Cell present in current but not baseline → new."""
        before = FIXTURES_DIR / "new_cell_before.xlsx"
        after  = FIXTURES_DIR / "new_cell_after.xlsx"
        drifts = _run_diff(before, after, ev)

        a3 = _drift_for(drifts, "Sheet", "A3")
        assert a3 is not None, "Expected drift on Sheet!A3"
        assert a3.kind == "new"
        assert a3.current_value == 99

    def test_deleted_cell(self, drift_fixtures, ev):
        """Cell in baseline but missing from current → deleted."""
        before = FIXTURES_DIR / "deleted_cell_before.xlsx"
        after  = FIXTURES_DIR / "deleted_cell_after.xlsx"
        drifts = _run_diff(before, after, ev)

        a3 = _drift_for(drifts, "Sheet", "A3")
        assert a3 is not None, "Expected drift on Sheet!A3"
        assert a3.kind == "deleted"
        assert a3.baseline_value == 3

    def test_exactly_one_drift_on_single_change(self, tmp_path, ev):
        """F2 AC: changing exactly one formula result → exactly one drift."""
        from openpyxl import Workbook

        # Before: three independent literal cells
        wb_before = Workbook()
        ws = wb_before.active
        ws.title = "Sheet"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = 30
        path_before = tmp_path / "one_change_before.xlsx"
        wb_before.save(path_before)

        # After: only A2 changes
        wb_after = Workbook()
        ws = wb_after.active
        ws.title = "Sheet"
        ws["A1"] = 10
        ws["A2"] = 999   # <── only this changed
        ws["A3"] = 30
        path_after = tmp_path / "one_change_after.xlsx"
        wb_after.save(path_after)

        drifts = _run_diff(path_before, path_after, ev)
        assert len(drifts) == 1, f"Expected exactly 1 drift, got {len(drifts)}: {drifts}"
        assert drifts[0].address == "A2"
        assert drifts[0].kind == "value_only"
        assert drifts[0].baseline_value == 20
        assert drifts[0].current_value  == 999

    def test_drift_has_sheet_address_before_after(self, tmp_path, ev):
        """F2 AC: drift report includes sheet, address, before, after."""
        from openpyxl import Workbook

        wb_b = Workbook(); ws = wb_b.active; ws.title = "Report"
        ws["C5"] = 500
        path_b = tmp_path / "meta_before.xlsx"; wb_b.save(path_b)

        wb_a = Workbook(); ws = wb_a.active; ws.title = "Report"
        ws["C5"] = 777
        path_a = tmp_path / "meta_after.xlsx"; wb_a.save(path_a)

        drifts = _run_diff(path_b, path_a, ev)
        assert len(drifts) == 1
        d = drifts[0]
        assert d.sheet   == "Report"
        assert d.address == "C5"
        assert d.baseline_value == 500
        assert d.current_value  == 777
