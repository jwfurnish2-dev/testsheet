"""
Generate M1 benchmark fixture workbooks with KNOWN expected values.

Three models of increasing formula complexity:
  model_simple.xlsx       — SUM, basic arithmetic, literal values
  model_intermediate.xlsx — IF, AVERAGE, MIN/MAX, cross-sheet refs
  model_complex.xlsx      — IFERROR, nested IF, chained cross-sheet, ROUND

Each file has a companion EXPECTED dict (returned by make_*) so the spike
script and tests can verify evaluator accuracy without hard-coding values
in two places.

Usage:
  python tests/fixtures/make_m1_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent


# ─────────────────────────────────────────────────────────────────────────────
# Model 1 — Simple
# Formulas: SUM, subtraction, multiplication, division
# All inputs are literals so expected values are trivially verifiable.
# ─────────────────────────────────────────────────────────────────────────────

def make_simple(path: Path) -> dict:
    """
    Sheet: Calc
      A1=10, A2=20, A3=30, A4=40
      B1=SUM(A1:A4)       → 100
      B2=B1*2             → 200
      B3=B1-A1            → 90
      B4=B2/B1            → 2.0
      B5=SUM(A1:A4)/4     → 25.0  (manual average)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"

    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["A4"] = 40
    ws["B1"] = "=SUM(A1:A4)"
    ws["B2"] = "=B1*2"
    ws["B3"] = "=B1-A1"
    ws["B4"] = "=B2/B1"
    ws["B5"] = "=SUM(A1:A4)/4"

    wb.save(path)

    expected = {
        "Calc": {
            "A1": 10, "A2": 20, "A3": 30, "A4": 40,
            "B1": 100,
            "B2": 200,
            "B3": 90,
            "B4": 2.0,
            "B5": 25.0,
        }
    }
    return expected


# ─────────────────────────────────────────────────────────────────────────────
# Model 2 — Intermediate
# Formulas: IF, AVERAGE, MIN, MAX, cross-sheet reference
# ─────────────────────────────────────────────────────────────────────────────

def make_intermediate(path: Path) -> dict:
    """
    Sheet: Data
      A1:A5 = [5, 15, 25, 35, 45]
      B1 = AVERAGE(A1:A5)   → 25.0
      B2 = MIN(A1:A5)       → 5
      B3 = MAX(A1:A5)       → 45
      B4 = IF(B1>20,"High","Low")  → "High"
      B5 = IF(B2<10,"Yes","No")    → "Yes"

    Sheet: Summary
      A1 = Data!B1          → 25.0  (cross-sheet)
      A2 = Data!B4          → "High"
      A3 = Data!B1 * 2      → 50.0
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    for i, v in enumerate([5, 15, 25, 35, 45], start=1):
        ws_data[f"A{i}"] = v

    ws_data["B1"] = "=AVERAGE(A1:A5)"
    ws_data["B2"] = "=MIN(A1:A5)"
    ws_data["B3"] = "=MAX(A1:A5)"
    ws_data["B4"] = '=IF(B1>20,"High","Low")'
    ws_data["B5"] = '=IF(B2<10,"Yes","No")'

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "=Data!B1"
    ws_summary["A2"] = "=Data!B4"
    ws_summary["A3"] = "=Data!B1*2"

    wb.save(path)

    expected = {
        "Data": {
            "A1": 5, "A2": 15, "A3": 25, "A4": 35, "A5": 45,
            "B1": 25.0,
            "B2": 5,
            "B3": 45,
            "B4": "High",
            "B5": "Yes",
        },
        "Summary": {
            "A1": 25.0,
            "A2": "High",
            "A3": 50.0,
        },
    }
    return expected


# ─────────────────────────────────────────────────────────────────────────────
# Model 3 — Complex
# Formulas: IFERROR, nested IF, ROUND, chained cross-sheet
# ─────────────────────────────────────────────────────────────────────────────

def make_complex(path: Path) -> dict:
    """
    Sheet: Inputs
      A1=100, A2=0, A3=250

    Sheet: Model
      B1 = IFERROR(Inputs!A1/Inputs!A2, "DIV_ERR")  → "DIV_ERR"  (div by zero caught)
      B2 = IFERROR(Inputs!A3/Inputs!A1, "DIV_ERR")  → 2.5
      B3 = ROUND(B2, 1)                              → 2.5
      B4 = IF(B2>2, IF(B2>3, "High", "Medium"), "Low")  → "Medium"
      B5 = Inputs!A1 + Inputs!A3                     → 350

    Sheet: Output
      C1 = Model!B2        → 2.5
      C2 = Model!B5 * 2   → 700
      C3 = ROUND(Model!B2 * Model!B5, 0)  → 875
    """
    wb = Workbook()

    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs["A1"] = 100
    ws_inputs["A2"] = 0
    ws_inputs["A3"] = 250

    ws_model = wb.create_sheet("Model")
    ws_model["B1"] = '=IFERROR(Inputs!A1/Inputs!A2,"DIV_ERR")'
    ws_model["B2"] = '=IFERROR(Inputs!A3/Inputs!A1,"DIV_ERR")'
    ws_model["B3"] = "=ROUND(B2,1)"
    ws_model["B4"] = '=IF(B2>2,IF(B2>3,"High","Medium"),"Low")'
    ws_model["B5"] = "=Inputs!A1+Inputs!A3"

    ws_output = wb.create_sheet("Output")
    ws_output["C1"] = "=Model!B2"
    ws_output["C2"] = "=Model!B5*2"
    ws_output["C3"] = "=ROUND(Model!B2*Model!B5,0)"

    wb.save(path)

    expected = {
        "Inputs": {"A1": 100, "A2": 0, "A3": 250},
        "Model": {
            "B1": "DIV_ERR",   # IFERROR caught div/0
            "B2": 2.5,
            "B3": 2.5,
            "B4": "Medium",
            "B5": 350,
        },
        "Output": {
            "C1": 2.5,
            "C2": 700,
            "C3": 875.0,
        },
    }
    return expected


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

ALL_MODELS = [
    ("model_simple.xlsx",       make_simple,       "Simple (SUM, arithmetic)"),
    ("model_intermediate.xlsx", make_intermediate, "Intermediate (IF, AVERAGE, cross-sheet)"),
    ("model_complex.xlsx",      make_complex,      "Complex (IFERROR, nested IF, ROUND, chained refs)"),
]


def build_all(out_dir: Path = FIXTURES_DIR) -> dict[str, dict]:
    """Build all 3 models and return {filename: expected_values}."""
    out_dir.mkdir(parents=True, exist_ok=True)
    results = {}
    for filename, builder, label in ALL_MODELS:
        p = out_dir / filename
        expected = builder(p)
        results[filename] = expected
        print(f"  Written: {p}  ({label})")
    return results


if __name__ == "__main__":
    print("Building M1 benchmark fixtures...")
    build_all()
    print("Done.")
