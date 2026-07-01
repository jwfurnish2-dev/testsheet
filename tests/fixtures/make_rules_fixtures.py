"""
Generate fixture workbooks for M3 rules engine tests.

One workbook covers all six rule types in a single "Rules" sheet,
making it easy to test pass/fail variants by tweaking one cell.

Sheet layout:
  A1:A5  = quarterly revenues (literals) — used for range_bound, monotonic
  B1     = SUM(A1:A5)                   — used for totals_tie
  C1     = AVERAGE(A1:A5)               — used for relationship
  D1:D5  = formula cells (=A1*1.1 etc.) — used for no_hardcode
  E1     = 0                             — used to trigger #DIV/0! in error tests
  F1     = =B1/E1  (or /1 for no-error) — used for no_error

Fixture workbooks produced:
  rules_pass.xlsx   — all rules should PASS
  rules_fail.xlsx   — all rules should FAIL (one violation per rule)
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent


def make_rules_pass(path: Path) -> None:
    """
    Workbook where every rule in rules_pass.yaml should PASS.

    Rules!A1:A5 = [100, 110, 120, 130, 140]  (increasing, all > 0)
    Rules!B1    = SUM(A1:A5) = 600
    Rules!C1    = AVERAGE(A1:A5) = 120
    Rules!D1:D5 = formulas (=A1, =A2, ...) — no literals
    Rules!E1    = 1   (non-zero denominator)
    Rules!F1    = =B1/E1 = 600  (no error)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"

    for i, v in enumerate([100, 110, 120, 130, 140], start=1):
        ws[f"A{i}"] = v

    ws["B1"] = "=SUM(A1:A5)"
    ws["C1"] = "=AVERAGE(A1:A5)"

    for i in range(1, 6):
        ws[f"D{i}"] = f"=A{i}"   # all formulas — passes no_hardcode

    ws["E1"] = 1                  # safe denominator
    ws["F1"] = "=B1/E1"          # evaluates cleanly — passes no_error

    wb.save(path)


def make_rules_fail(path: Path) -> None:
    """
    Workbook where every rule in rules_fail.yaml should FAIL.

    Rules!A1:A5 = [100, 50, 120, 130, 140]  (A2=50 breaks monotonic + range_bound if min=80)
    Rules!B1    = 999  (literal override — breaks totals_tie)
    Rules!C1    = =AVERAGE(A1:A5)
    Rules!D1:D5 = D3 is a literal (42) — breaks no_hardcode
    Rules!E1    = 0                          (zero denominator)
    Rules!F1    = =B1/E1                     (→ #DIV/0! — breaks no_error)
    Rules!G1    = -50                        (breaks relationship G1 >= 0)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"

    for i, v in enumerate([100, 50, 120, 130, 140], start=1):
        ws[f"A{i}"] = v           # A2=50 breaks monotonic

    ws["B1"] = 999                # literal, NOT =SUM(A1:A5) → breaks totals_tie
    ws["C1"] = "=AVERAGE(A1:A5)"

    for i in range(1, 6):
        if i == 3:
            ws[f"D{i}"] = 42      # hardcoded literal — breaks no_hardcode
        else:
            ws[f"D{i}"] = f"=A{i}"

    ws["E1"] = 0                  # zero denominator
    ws["F1"] = "=B1/E1"          # → #DIV/0! — breaks no_error
    ws["G1"] = -50                # breaks relationship

    wb.save(path)


def make_rules_yaml(path: Path, workbook_name: str, fail: bool = False) -> None:
    """Write a rules.yaml targeting the Rules sheet."""
    rules = [
        {
            "id": "revenues_positive",
            "type": "range_bound",
            "range": "Rules!A1:A5",
            "min": 80,            # passes for [100,110,120,130,140]; fails for A2=50
        },
        {
            "id": "no_formula_errors",
            "type": "no_error",
            "range": "Rules!F1",  # passes when E1≠0; fails when E1=0
        },
        {
            "id": "no_hardcoded_drivers",
            "type": "no_hardcode_in_range",
            "range": "Rules!D1:D5",  # passes when all =formula; fails when D3=42
        },
        {
            "id": "total_ties_sum",
            "type": "totals_tie",
            "total_cell": "Rules!B1",
            "sum_range": "Rules!A1:A5",
            "tolerance": 0.01,    # passes when B1=SUM; fails when B1=999
        },
        {
            "id": "revenues_increasing",
            "type": "monotonic",
            "range": "Rules!A1:A5",
            "direction": "increasing",  # fails when A2=50
        },
        {
            "id": "gross_profit_nonneg",
            "type": "relationship",
            "expression": "Rules!G1 >= 0" if fail else "Rules!C1 >= 0",
            # fail version: G1=-50 → False; pass version: C1=avg ≥ 0 → True
        },
    ]
    data = {"rules": rules}
    path.write_text(yaml.dump(data, sort_keys=False), encoding="utf-8")


def build_all_rules(out_dir: Path = FIXTURES_DIR) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    make_rules_pass(out_dir / "rules_pass.xlsx")
    print("  Written: rules_pass.xlsx")

    make_rules_fail(out_dir / "rules_fail.xlsx")
    print("  Written: rules_fail.xlsx")

    make_rules_yaml(out_dir / "rules_pass.yaml", "rules_pass.xlsx", fail=False)
    print("  Written: rules_pass.yaml")

    make_rules_yaml(out_dir / "rules_fail.yaml", "rules_fail.xlsx", fail=True)
    print("  Written: rules_fail.yaml")


if __name__ == "__main__":
    print("Building rules fixtures...")
    build_all_rules()
    print("Done.")
