"""
Generate fixture workbooks for the test suite.

Run once (or via pytest conftest setup) to produce:
  tests/fixtures/good_model.xlsx   — "known-good" FP&A-style model
  tests/fixtures/broken_model.xlsx — same model with a planted formula error

Usage:
  python tests/fixtures/make_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook


FIXTURES_DIR = Path(__file__).parent


def make_good_model(path: Path) -> None:
    """
    A minimal 3-sheet FP&A model:
      Model   — revenue & cost rows, quarterly columns
      Summary — totals pulling from Model
      Output  — final KPI cells
    """
    wb = Workbook()

    # ── Model sheet ────────────────────────────────────────────────────────
    ws_model = wb.active
    ws_model.title = "Model"

    headers = ["", "Q1", "Q2", "Q3", "Q4", "Total"]
    ws_model.append(headers)

    revenues = [100, 110, 120, 130]
    costs    = [60,  65,  70,  75]

    ws_model.append(["Revenue"] + revenues + [f"=SUM(B2:E2)"])
    ws_model.append(["Cost"]    + costs    + [f"=SUM(B3:E3)"])
    ws_model.append(["Gross Profit",
                     "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3",
                     "=SUM(B4:E4)"])

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Metric"
    ws_summary["B1"] = "Value"
    ws_summary["A2"] = "Total Revenue"
    ws_summary["B2"] = "=Model!F2"
    ws_summary["A3"] = "Total Cost"
    ws_summary["B3"] = "=Model!F3"
    ws_summary["A4"] = "Gross Profit"
    ws_summary["B4"] = "=Model!F4"
    ws_summary["A5"] = "Gross Margin %"
    ws_summary["B5"] = "=B4/B2"

    # ── Output sheet ──────────────────────────────────────────────────────
    ws_output = wb.create_sheet("Output")
    ws_output["A1"] = "Net Income (stub)"
    ws_output["B1"] = "=Summary!B4"  # gross profit as proxy

    wb.save(path)
    print(f"Written: {path}")


def make_broken_model(path: Path) -> None:
    """
    Same structure as good_model but with two planted errors:
      1. Revenue Q3 changed from 120 → 999 (value drift)
      2. Summary!B5 formula broken → introduces #DIV/0!
    """
    wb = Workbook()

    ws_model = wb.active
    ws_model.title = "Model"

    headers = ["", "Q1", "Q2", "Q3", "Q4", "Total"]
    ws_model.append(headers)

    revenues = [100, 110, 999, 130]   # <── planted change: 120 → 999
    costs    = [60,  65,  70,  75]

    ws_model.append(["Revenue"] + revenues + [f"=SUM(B2:E2)"])
    ws_model.append(["Cost"]    + costs    + [f"=SUM(B3:E3)"])
    ws_model.append(["Gross Profit",
                     "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3",
                     "=SUM(B4:E4)"])

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Metric"
    ws_summary["B1"] = "Value"
    ws_summary["A2"] = "Total Revenue"
    ws_summary["B2"] = "=Model!F2"
    ws_summary["A3"] = "Total Cost"
    ws_summary["B3"] = "=Model!F3"
    ws_summary["A4"] = "Gross Profit"
    ws_summary["B4"] = "=Model!F4"
    ws_summary["A5"] = "Gross Margin %"
    ws_summary["B5"] = "=B4/0"     # <── planted error: #DIV/0!

    ws_output = wb.create_sheet("Output")
    ws_output["A1"] = "Net Income (stub)"
    ws_output["B1"] = "=Summary!B4"

    wb.save(path)
    print(f"Written: {path}")


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    make_good_model(FIXTURES_DIR / "good_model.xlsx")
    make_broken_model(FIXTURES_DIR / "broken_model.xlsx")
    print("Fixtures ready.")
