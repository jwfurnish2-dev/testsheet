"""
M3 rules engine tests — F3 acceptance criteria.

AC: Each of the six rule types has a passing and a failing fixture test.

Rule types covered:
  range_bound          — min/max on a cell/range
  no_error             — no Excel error string in range
  no_hardcode_in_range — every cell in range must be a formula
  totals_tie           — total_cell == SUM(sum_range) within tolerance
  monotonic            — values in range are monotonically increasing/decreasing
  relationship         — arbitrary cross-cell expression evaluates to True
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
import yaml

FIXTURES_DIR = Path(__file__).parent / "fixtures"
sys.path.insert(0, str(FIXTURES_DIR))

# Skip entire module if formulas not installed
pytest.importorskip("formulas", reason="formulas not installed — run: pip install formulas")


# ── Module fixtures ───────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def ev():
    from testsheet.evaluator.formulas_eval import FormulasEvaluator
    return FormulasEvaluator()


@pytest.fixture(scope="module")
def rule_fixtures():
    """Build pass/fail workbooks and rule yaml files."""
    from make_rules_fixtures import build_all_rules
    build_all_rules(FIXTURES_DIR)
    return {
        "pass_wb":   FIXTURES_DIR / "rules_pass.xlsx",
        "fail_wb":   FIXTURES_DIR / "rules_fail.xlsx",
        "pass_yaml": FIXTURES_DIR / "rules_pass.yaml",
        "fail_yaml": FIXTURES_DIR / "rules_fail.yaml",
    }


# ── Helpers ───────────────────────────────────────────────────────────────────

def _eval_rules(workbook_path: Path, rules_yaml: Path, ev):
    """Load rules.yaml, evaluate against workbook, return list of RuleResult."""
    from testsheet.rules.engine import load_rules, evaluate_rules
    rules = load_rules(rules_yaml)
    return evaluate_rules(workbook_path, rules, evaluator=ev)


def _result_for(results, rule_id: str):
    matches = [r for r in results if r.rule_id == rule_id]
    return matches[0] if matches else None


# ── range_bound ───────────────────────────────────────────────────────────────

class TestRangeBound:
    def test_pass(self, rule_fixtures, ev):
        """All revenues ≥ 80 → range_bound passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "revenues_positive")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """A2=50 < min=80 → range_bound fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "revenues_positive")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"
        assert "50" in r.message or "A2" in r.message

    def test_unit_pass(self):
        from testsheet.rules.engine import _rule_range_bound
        computed = {"Sheet": {"B1": 150.0}}
        rule = {"id": "r", "type": "range_bound", "cell": "Sheet!B1", "min": 0, "max": 200}
        assert _rule_range_bound(rule, computed, None).passed

    def test_unit_fail_below_min(self):
        from testsheet.rules.engine import _rule_range_bound
        computed = {"Sheet": {"B1": -5.0}}
        rule = {"id": "r", "type": "range_bound", "cell": "Sheet!B1", "min": 0}
        assert not _rule_range_bound(rule, computed, None).passed

    def test_unit_fail_above_max(self):
        from testsheet.rules.engine import _rule_range_bound
        computed = {"Sheet": {"B1": 999.0}}
        rule = {"id": "r", "type": "range_bound", "cell": "Sheet!B1", "max": 100}
        assert not _rule_range_bound(rule, computed, None).passed


# ── no_error ──────────────────────────────────────────────────────────────────

class TestNoError:
    def test_pass(self, rule_fixtures, ev):
        """F1=B1/1 produces clean value → no_error passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "no_formula_errors")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """F1=B1/0 → #DIV/0! → no_error fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "no_formula_errors")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"

    def test_unit_pass(self):
        from testsheet.rules.engine import _rule_no_error
        computed = {"Sheet": {"A1": 42, "A2": "hello"}}
        rule = {"id": "r", "type": "no_error", "range": "Sheet!A1:A2"}
        assert _rule_no_error(rule, computed, None).passed

    def test_unit_fail(self):
        from testsheet.rules.engine import _rule_no_error
        computed = {"Sheet": {"A1": "#REF!"}}
        rule = {"id": "r", "type": "no_error", "range": "Sheet!A1"}
        assert not _rule_no_error(rule, computed, None).passed

    def test_all_error_types_caught(self):
        from testsheet.rules.engine import _rule_no_error
        for err in ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"]:
            computed = {"S": {"A1": err}}
            rule = {"id": "r", "type": "no_error", "range": "S!A1"}
            result = _rule_no_error(rule, computed, None)
            assert not result.passed, f"Expected {err} to fail no_error"


# ── no_hardcode_in_range ──────────────────────────────────────────────────────

class TestNoHardcode:
    def test_pass(self, rule_fixtures, ev):
        """D1:D5 all formulas → no_hardcode passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "no_hardcoded_drivers")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """D3=42 (literal) → no_hardcode fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "no_hardcoded_drivers")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"
        assert "D3" in r.message

    def test_unit_pass(self, tmp_path):
        """All formula cells → passes."""
        from openpyxl import Workbook
        import openpyxl
        from testsheet.rules.engine import _rule_no_hardcode

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "=1+1"
        ws["A2"] = "=A1*2"
        p = tmp_path / "t.xlsx"; wb.save(p)
        wb_f = openpyxl.load_workbook(p, data_only=False, read_only=True)

        rule = {"id": "r", "type": "no_hardcode_in_range", "range": "S!A1:A2"}
        assert _rule_no_hardcode(rule, {}, wb_f).passed
        wb_f.close()

    def test_unit_fail(self, tmp_path):
        """Literal in range → fails."""
        from openpyxl import Workbook
        import openpyxl
        from testsheet.rules.engine import _rule_no_hardcode

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "=1+1"
        ws["A2"] = 99          # hardcoded
        p = tmp_path / "t.xlsx"; wb.save(p)
        wb_f = openpyxl.load_workbook(p, data_only=False, read_only=True)

        rule = {"id": "r", "type": "no_hardcode_in_range", "range": "S!A1:A2"}
        result = _rule_no_hardcode(rule, {}, wb_f)
        assert not result.passed
        assert "A2" in result.message
        wb_f.close()


# ── totals_tie ────────────────────────────────────────────────────────────────

class TestTotalsTie:
    def test_pass(self, rule_fixtures, ev):
        """B1=SUM(A1:A5) → totals_tie passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "total_ties_sum")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """B1=999 (literal, doesn't match SUM) → totals_tie fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "total_ties_sum")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"

    def test_unit_pass_exact(self):
        from testsheet.rules.engine import _rule_totals_tie
        computed = {"S": {"B1": 100.0, "A1": 40.0, "A2": 60.0}}
        rule = {"id": "r", "type": "totals_tie",
                "total_cell": "S!B1", "sum_range": "S!A1:A2", "tolerance": 0.01}
        assert _rule_totals_tie(rule, computed, None).passed

    def test_unit_pass_within_tolerance(self):
        from testsheet.rules.engine import _rule_totals_tie
        computed = {"S": {"B1": 100.005, "A1": 40.0, "A2": 60.0}}
        rule = {"id": "r", "type": "totals_tie",
                "total_cell": "S!B1", "sum_range": "S!A1:A2", "tolerance": 0.01}
        assert _rule_totals_tie(rule, computed, None).passed

    def test_unit_fail(self):
        from testsheet.rules.engine import _rule_totals_tie
        computed = {"S": {"B1": 999.0, "A1": 40.0, "A2": 60.0}}
        rule = {"id": "r", "type": "totals_tie",
                "total_cell": "S!B1", "sum_range": "S!A1:A2", "tolerance": 0.01}
        assert not _rule_totals_tie(rule, computed, None).passed


# ── monotonic ─────────────────────────────────────────────────────────────────

class TestMonotonic:
    def test_pass(self, rule_fixtures, ev):
        """[100,110,120,130,140] increasing → monotonic passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "revenues_increasing")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """[100,50,120,130,140] — A2 dips → monotonic fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "revenues_increasing")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"

    def test_unit_increasing_pass(self):
        from testsheet.rules.engine import _rule_monotonic
        computed = {"S": {"A1": 1, "A2": 2, "A3": 3}}
        rule = {"id": "r", "type": "monotonic", "range": "S!A1:A3",
                "direction": "increasing"}
        assert _rule_monotonic(rule, computed, None).passed

    def test_unit_increasing_fail(self):
        from testsheet.rules.engine import _rule_monotonic
        computed = {"S": {"A1": 1, "A2": 5, "A3": 3}}  # A3 dips
        rule = {"id": "r", "type": "monotonic", "range": "S!A1:A3",
                "direction": "increasing"}
        assert not _rule_monotonic(rule, computed, None).passed

    def test_unit_decreasing_pass(self):
        from testsheet.rules.engine import _rule_monotonic
        computed = {"S": {"A1": 10, "A2": 7, "A3": 3}}
        rule = {"id": "r", "type": "monotonic", "range": "S!A1:A3",
                "direction": "decreasing"}
        assert _rule_monotonic(rule, computed, None).passed

    def test_unit_flat_non_strict_pass(self):
        """Equal adjacent values pass non-strict monotonic."""
        from testsheet.rules.engine import _rule_monotonic
        computed = {"S": {"A1": 5, "A2": 5, "A3": 6}}
        rule = {"id": "r", "type": "monotonic", "range": "S!A1:A3",
                "direction": "increasing", "strict": False}
        assert _rule_monotonic(rule, computed, None).passed


# ── relationship ──────────────────────────────────────────────────────────────

class TestRelationship:
    def test_pass(self, rule_fixtures, ev):
        """C1=AVERAGE ≥ 0 → relationship passes."""
        results = _eval_rules(rule_fixtures["pass_wb"], rule_fixtures["pass_yaml"], ev)
        r = _result_for(results, "gross_profit_nonneg")
        assert r is not None
        assert r.passed, f"Expected PASS: {r.message}"

    def test_fail(self, rule_fixtures, ev):
        """G1=-50 ≥ 0 → relationship fails."""
        results = _eval_rules(rule_fixtures["fail_wb"], rule_fixtures["fail_yaml"], ev)
        r = _result_for(results, "gross_profit_nonneg")
        assert r is not None
        assert not r.passed, "Expected FAIL but got PASS"

    def test_unit_pass(self):
        from testsheet.rules.engine import _rule_relationship
        computed = {"Sheet": {"B4": 100.0, "B2": 400.0}}
        rule = {"id": "r", "type": "relationship",
                "expression": "Sheet!B4 >= Sheet!B2 * 0.1"}
        assert _rule_relationship(rule, computed, None).passed

    def test_unit_fail(self):
        from testsheet.rules.engine import _rule_relationship
        computed = {"Sheet": {"B4": -10.0}}
        rule = {"id": "r", "type": "relationship",
                "expression": "Sheet!B4 >= 0"}
        assert not _rule_relationship(rule, computed, None).passed

    def test_unit_cross_sheet(self):
        from testsheet.rules.engine import _rule_relationship
        computed = {"Summary": {"B2": 500.0}, "Detail": {"C10": 490.0}}
        rule = {"id": "r", "type": "relationship",
                "expression": "Summary!B2 >= Detail!C10"}
        assert _rule_relationship(rule, computed, None).passed
