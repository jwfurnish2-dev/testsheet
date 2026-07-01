"""
M5 hardening tests.

Covers:
  - Structural-change heuristic (detect_structural_change)
  - Float tolerance config (load_config + diff integration)
  - Config loading (defaults, overrides, missing file)
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest
import yaml

FIXTURES_DIR = Path(__file__).parent / "fixtures"
sys.path.insert(0, str(FIXTURES_DIR))

pytest.importorskip("formulas", reason="formulas not installed — run: pip install formulas")


# ── Config loader ─────────────────────────────────────────────────────────────

class TestLoadConfig:

    def test_defaults_when_no_file(self, tmp_path):
        from testsheet.config import load_config
        cfg = load_config(tmp_path)
        assert cfg["diff"]["rel_tol"] == pytest.approx(1e-9)
        assert cfg["diff"]["abs_tol"] == pytest.approx(1e-12)
        assert cfg["structural_change_threshold"] == pytest.approx(0.5)

    def test_override_rel_tol(self, tmp_path):
        from testsheet.config import load_config
        (tmp_path / "config.yaml").write_text(
            yaml.dump({"diff": {"rel_tol": 1e-4}}), encoding="utf-8"
        )
        cfg = load_config(tmp_path)
        assert cfg["diff"]["rel_tol"] == pytest.approx(1e-4)
        assert cfg["diff"]["abs_tol"] == pytest.approx(1e-12)  # default unchanged

    def test_override_abs_tol(self, tmp_path):
        from testsheet.config import load_config
        (tmp_path / "config.yaml").write_text(
            yaml.dump({"diff": {"abs_tol": 0.001}}), encoding="utf-8"
        )
        cfg = load_config(tmp_path)
        assert cfg["diff"]["abs_tol"] == pytest.approx(0.001)

    def test_override_structural_threshold(self, tmp_path):
        from testsheet.config import load_config
        (tmp_path / "config.yaml").write_text(
            yaml.dump({"structural_change_threshold": 0.25}), encoding="utf-8"
        )
        cfg = load_config(tmp_path)
        assert cfg["structural_change_threshold"] == pytest.approx(0.25)

    def test_partial_override_preserves_defaults(self, tmp_path):
        from testsheet.config import load_config
        (tmp_path / "config.yaml").write_text(
            yaml.dump({"structural_change_threshold": 0.1}), encoding="utf-8"
        )
        cfg = load_config(tmp_path)
        assert cfg["diff"]["rel_tol"] == pytest.approx(1e-9)  # default
        assert cfg["structural_change_threshold"] == pytest.approx(0.1)

    def test_empty_config_file_uses_defaults(self, tmp_path):
        from testsheet.config import load_config
        (tmp_path / "config.yaml").write_text("", encoding="utf-8")
        cfg = load_config(tmp_path)
        assert cfg["diff"]["rel_tol"] == pytest.approx(1e-9)


# ── Structural-change heuristic ───────────────────────────────────────────────

class TestDetectStructuralChange:

    def _make_baseline(self, n_cells: int) -> dict:
        """Baseline with n_cells in a single sheet."""
        cells = {f"A{i}": {"value": i, "formula": None, "hash": ""} for i in range(1, n_cells + 1)}
        return {"sheets": {"Sheet": cells}}

    def _make_drifts(self, n: int):
        from testsheet.diff import Drift
        return [
            Drift("Sheet", f"A{i}", "value_only", i, i + 1, None, None)
            for i in range(1, n + 1)
        ]

    def test_no_warning_below_threshold(self):
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(4)  # 40% < 50%
        assert detect_structural_change(drifts, baseline, threshold=0.5) is None

    def test_warning_above_threshold(self):
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(6)  # 60% > 50%
        result = detect_structural_change(drifts, baseline, threshold=0.5)
        assert result is not None
        assert result.drift_count == 6
        assert result.baseline_cell_count == 10
        assert result.drift_fraction == pytest.approx(0.6)

    def test_warning_at_exact_threshold_excluded(self):
        """Exactly at threshold → no warning (strictly greater than)."""
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(5)  # exactly 50%
        assert detect_structural_change(drifts, baseline, threshold=0.5) is None

    def test_custom_threshold(self):
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(3)  # 30% > 25% threshold
        result = detect_structural_change(drifts, baseline, threshold=0.25)
        assert result is not None

    def test_empty_baseline_no_warning(self):
        """Empty baseline → don't divide by zero."""
        from testsheet.diff import detect_structural_change
        result = detect_structural_change([], {"sheets": {}}, threshold=0.5)
        assert result is None

    def test_message_contains_percentage(self):
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(7)
        result = detect_structural_change(drifts, baseline, threshold=0.5)
        assert "70%" in result.message

    def test_as_dict(self):
        from testsheet.diff import detect_structural_change
        baseline = self._make_baseline(10)
        drifts = self._make_drifts(8)
        result = detect_structural_change(drifts, baseline, threshold=0.5)
        d = result.as_dict()
        for key in ("drift_count", "baseline_cell_count", "drift_fraction", "threshold", "message"):
            assert key in d


# ── Float tolerance integration ───────────────────────────────────────────────

class TestFloatToleranceIntegration:

    def test_tight_tolerance_flags_small_diff(self, tmp_path):
        """With rel_tol=0, even a tiny float difference is flagged."""
        from testsheet.diff import diff_workbook, _values_equal
        # Unit test — _values_equal respects tolerances
        assert not _values_equal(1.0, 1.0 + 1e-8, rel_tol=0.0, abs_tol=0.0)

    def test_loose_tolerance_ignores_small_diff(self):
        from testsheet.diff import _values_equal
        assert _values_equal(1.0, 1.0 + 1e-8, rel_tol=1e-6, abs_tol=1e-6)

    def test_config_rel_tol_flows_to_diff(self, tmp_path):
        """
        E2E: config.yaml with rel_tol=0.1 means values within 10% are equal.
        Capture baseline with value 100, run with value 105 (5% diff) → no drift.
        """
        from typer.testing import CliRunner
        from testsheet.cli import app

        # Build a workbook with value 100
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = 100.0
        wb_path = tmp_path / "model.xlsx"
        wb.save(wb_path)

        runner = CliRunner()

        # Baseline
        r = runner.invoke(app, ["baseline", str(wb_path)])
        assert r.exit_code == 0, r.output

        # Write a loose tolerance config
        ts_dir = tmp_path / ".testsheet"
        (ts_dir / "config.yaml").write_text(
            yaml.dump({"diff": {"rel_tol": 0.1, "abs_tol": 0.1}}),
            encoding="utf-8",
        )

        # Modify workbook: A1 = 105 (5% drift, within 10% tolerance)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Data"
        ws2["A1"] = 105.0
        wb2.save(wb_path)

        r = runner.invoke(app, ["run", str(wb_path)])
        assert r.exit_code == 0, f"Expected no drift with loose tolerance:\n{r.output}"

    def test_config_rel_tol_strict_catches_diff(self, tmp_path):
        """
        With default rel_tol=1e-9, a 5% difference IS flagged.
        """
        from typer.testing import CliRunner
        from testsheet.cli import app
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = 100.0
        wb_path = tmp_path / "model.xlsx"
        wb.save(wb_path)

        runner = CliRunner()
        r = runner.invoke(app, ["baseline", str(wb_path)])
        assert r.exit_code == 0

        # No config.yaml → defaults (rel_tol=1e-9)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Data"
        ws2["A1"] = 105.0
        wb2.save(wb_path)

        r = runner.invoke(app, ["run", str(wb_path)])
        assert r.exit_code == 1, f"Expected drift with tight tolerance:\n{r.output}"


# ── Structural change in CLI output ──────────────────────────────────────────

class TestStructuralChangeCliOutput:

    def test_structural_warning_in_json_report(self, tmp_path):
        """
        When >50% of cells drift the JSON report contains structural_change key.
        """
        import json
        from typer.testing import CliRunner
        from testsheet.cli import app
        from openpyxl import Workbook

        # Workbook with 4 cells
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for i in range(1, 5):
            ws[f"A{i}"] = i * 10.0
        wb_path = tmp_path / "m.xlsx"
        wb.save(wb_path)

        runner = CliRunner()
        runner.invoke(app, ["baseline", str(wb_path)])

        # Change 3 of 4 cells → 75% drift (> 50% threshold)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "S"
        ws2["A1"] = 10.0         # unchanged
        ws2["A2"] = 999.0        # changed
        ws2["A3"] = 999.0        # changed
        ws2["A4"] = 999.0        # changed
        wb2.save(wb_path)

        json_out = tmp_path / "r.json"
        runner.invoke(app, ["run", str(wb_path), "--json", str(json_out), "--no-fail-on-drift"])
        data = json.loads(json_out.read_text())
        assert "structural_change" in data
        assert data["structural_change"] is not None
        assert data["structural_change"]["drift_fraction"] > 0.5

    def test_no_structural_warning_on_small_drift(self, tmp_path):
        import json
        from typer.testing import CliRunner
        from testsheet.cli import app
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for i in range(1, 11):
            ws[f"A{i}"] = i * 10.0
        wb_path = tmp_path / "m.xlsx"
        wb.save(wb_path)

        runner = CliRunner()
        runner.invoke(app, ["baseline", str(wb_path)])

        # Change only 1 of 10 cells → 10% drift (< 50%)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "S"
        for i in range(1, 11):
            ws2[f"A{i}"] = i * 10.0
        ws2["A1"] = 999.0
        wb2.save(wb_path)

        json_out = tmp_path / "r.json"
        runner.invoke(app, ["run", str(wb_path), "--json", str(json_out), "--no-fail-on-drift"])
        data = json.loads(json_out.read_text())
        assert data.get("structural_change") is None
